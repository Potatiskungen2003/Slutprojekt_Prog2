
from getpwd import getpwd
import openpyxl
import time

from openpyxl import Workbook

class Bank:
    def __init__(self, name, password, balance, role):
        self.name = name
        self.password = password
        self.balance = balance
        self.role = role

    def __str__(self):
        return (
            "----------------------------------------"
            f"\nVälkommen!\nInloggad användare: {self.name}"
        )
    
    def roles(self):
        return self.role

    def account_balance(self):
        return (
            "\n------------------------------"
            f"\nDitt saldo är: {self.balance}"
            " kr\n------------------------------"
        )

    def profile(self):
        return (
            "\n------------------------------"
            "\nMin profil:\n"
            f"\nNamn: {self.name}\nLösenord: {self.password}"
            f"\nKontotyp: {self.role}"
            "\n------------------------------"
        )

    def donate(self):
            try:
                user_donation = int(input("\nHur mycket pengar(kr) vill du donera till banken?\n: "))
                if user_donation <= 0:
                    print("\nDu kan inte donera mindre än 1 krona.")
                elif user_donation < self.balance:
                    self.balance = self.balance - user_donation
                    print("\nTack för din donation!\nVi älskar dina pengar!")
                elif user_donation > self.balance:
                    print("Betalning misslyckades :(\nDitt saldo är för lågt!\nDu kan se ditt ""saldo"" under Mitt saldo.")
                
            except:
                print("Du kan bara donera hela kronor.")
            print("----------------------------------------")

    def save_data(self):
        path = "./Bank_info.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        max_ro = sheet_obj.max_row

        for x in range(1, max_ro + 1):

            name = sheet_obj.cell(row = x, column = 1)
            if name.value == self.name:
                sheet_obj.cell(row=x, column=1).value = self.name
                sheet_obj.cell(row=x, column=2).value = self.password
                sheet_obj.cell(row=x, column=3).value = self.balance
                print()
                wb_obj.save("./Bank_info.xlsx")
                break # Saker inte funka som det ska

class Admin(Bank):
    def __init__(self, name, password, balance, role):
        super().__init__(name, password, balance, role)

    def see_accounts():
        path = "./Bank_info.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        max_ro = sheet_obj.max_row
        
        list = ["Anv.namn: ", "Lösenord: ", "Saldo(kr): ", "Kontotyp: "]

        for account in range(1, max_ro + 1):
            print(f"\n--------------------\nKonto nr.{account}")
            column_counter = 1
            print()
            for word in list:
                print(f"{word}{sheet_obj.cell(row = account, column = column_counter).value}")
                column_counter += 1
                time.sleep(0.1)
            time.sleep(0.5)
        print("\n--------------------")

class Adult(Bank):
    def __init__(self, name, password, balance, role):
        super().__init__(name, password, balance, role)

class Minor(Bank):
    def __init__(self, name, password, balance, role):
        super().__init__(name, password, balance, role)

    def donate(self):
        print("Detta konto är inställt som underårig."
        "\nFråga dina föräldrar om hjälp")

# Functions
def login():
    path = "./Bank_info.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_ro = sheet_obj.max_row
    
    trys = 4
    run = True
    while run == True:
        
        input_name = str(input("\n\nAnvändarnamn: "))
        password = getpwd("Lösenord: ")

        for x in range(1, max_ro + 1):

            name = sheet_obj.cell(row = x, column = 1)
            
            pass_check = sheet_obj.cell(row = x, column = 2)

            if password == pass_check.value:
                if name.value == input_name:
                    print("\nDu hade rätt användarnamn och lösenord")
                    run = False
                    break

        if password != pass_check.value or name.value != input_name:
                trys = trys-1
                print("Fel användarnamn eller lösenord")
                print("Du har", trys, "försök kvar")
                if trys < 1:
                    print("Du kan logga in igen om:")
                    for i in range(6):
                        print(f"{(6-i)*10} sekunder")
                        time.sleep(1)
                    trys = 4


    role = sheet_obj.cell(row = x, column = 4)
    balance = sheet_obj.cell(row = x, column = 3)
    if role.value == "Admin":
        user = Admin(name.value, pass_check.value, balance.value, role.value)
    
    elif role.value == "Adult":
        user = Adult(name.value, pass_check.value, balance.value, role.value)
    
    elif role.value == "Minor":
        user = Minor(name.value, pass_check.value, balance.value, role.value)
    
    print(user)
    return user

def donate(user = ""):
    print("\n----------------------------------------")
    while True:
        try:
            user_agree = str(input("Vill du donera till banken?\nja/nej: "))
            user_agree = user_agree.lower()
        except:
            print("\nHoppsan! Något gick fel.\n")
            time.sleep(2)
        if user_agree == "ja":
            user.donate()
            break
        elif user_agree == "nej":
            print(
                "\nOkej :(\nTänk bara på allt kaffe "
            "vi hade kunnat \nköpa med dina pengar."
            "\n----------------------------------------"
            )
            break
        else:
            print("\nHoppsan! Något gick fel.\n")
            time.sleep(2)


def main():
    print("\nVälkommen till Bankens internettjänst!")
    user = login()
    loginacces = True
    while loginacces == True:
        role = user.roles()
        print("\nVad vill du göra?"
            "\n(1) - Mitt saldo\n(2) - Min profil"
            "\n(3) - Byt användare")
        if role == "Admin":
            print("(4) - Kundöversikt")
        else:
            print("(4) - Donera till banken")
        print("(5) - Logga ut och avsluta")
        try:
            user_input = int(input("Nr: "))
        except:
            print("\nAnvänd hela siffror.")
            time.sleep(2)
        try:
            if user_input == 1:
                print(user.account_balance())# Hämta saldo
            elif user_input == 2:
                print(user.profile())# Gå till profil
            elif user_input == 3:
                user.save_data()
                user = login()# Byta användare
            elif user_input == 4:
                if role == "Admin":
                    Admin.see_accounts()
                else:
                    donate(user)# Donera pengar till banken
            elif user_input == 5:
                user.save_data()# Avsluta och spara data
                print("Välkommen åter!")
                loginacces = False
            else:
                print("\nAnvänd siffror 1-5")
                time.sleep(2)
        except:
            pass
main()