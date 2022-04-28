# Python program to read an excel file
 
# import openpyxl module
import openpyxl
 
# Give the location of the file
path = "C:/Users/edward.wallbeckhall/Programmering 2/Slutprojekt_Prog2/Bank_info.xlsx"
 
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
 
# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0

max_col = sheet_obj.max_column
max_ro = sheet_obj.max_row


for x in range(1, max_ro + 1):
    for i in range(1, max_col + 1):
        # Cell object is created by using
        # sheet object's cell() method.
        cell_obj = sheet_obj.cell(row = x, column = i)
        
        # Print value of cell object
        # using the value attribute
        print(cell_obj.value, end=' ')
    print()



anv_nam = str(input('Vae heteru: '))

for x in range(1, max_ro + 1):

    anv = sheet_obj.cell(row = x, column = 1)

    if anv.value == anv_nam:
        password = str(input('Password: '))
        pass_check = sheet_obj.cell(row = x, column = 2)

        if password == pass_check.value:
            print('Du hade rätt lösenord')
            break
    

